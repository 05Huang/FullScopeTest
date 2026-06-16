"""
报告导出服务

支持 PDF 和 Excel 两种格式的测试报告导出。
PDF 使用 ReportLab 生成，Excel 使用 openpyxl 生成。

功能：
- PDF 报告：项目信息、执行摘要、通过率、失败用例详情、执行时间
- Excel 报告：汇总 Sheet + 用例详情 Sheet + 性能指标 Sheet
- 支持自定义报告范围（按时间/按项目/按测试类型）
"""
import io
import os
from datetime import datetime, timedelta
from typing import Optional
from ..extensions import db
from ..models.test_run import TestRun
from ..models.project import Project
from ..utils.exceptions import NotFoundError, ValidationError
from ..core.logging import get_logger

logger = get_logger(__name__)


def generate_pdf_report(test_run_id: int) -> Optional[bytes]:
    """
    生成 PDF 格式的测试报告

    Args:
        test_run_id: TestRun ID

    Returns:
        PDF 文件的 bytes，None 表示 ReportLab 未安装
    """
    run = TestRun.query.get(test_run_id)
    if not run:
        raise NotFoundError("测试执行记录", test_run_id)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        logger.warning("ReportLab 未安装，无法生成 PDF")
        return None

    project = Project.query.get(run.project_id)

    # 尝试注册中文字体（可选）
    _try_register_chinese_font()

    # 构建 PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            topMargin=20 * mm, bottomMargin=20 * mm)

    styles = getSampleStyleSheet()
    elements = []

    # 标题
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(f'Test Report #{run.id}', title_style))
    elements.append(Spacer(1, 6 * mm))

    # 项目信息
    elements.append(Paragraph('Project Information', styles['Heading2']))
    project_info = [
        ['Project', project.name if project else f'ID: {run.project_id}'],
        ['Test Type', run.test_type or 'N/A'],
        ['Test Object', run.test_object_name or 'N/A'],
        ['Status', run.status or 'N/A'],
        ['Triggered By', run.triggered_by or 'manual'],
    ]
    elements.append(_build_table(project_info, col_widths=[120, 300]))
    elements.append(Spacer(1, 6 * mm))

    # 执行摘要
    elements.append(Paragraph('Execution Summary', styles['Heading2']))
    pass_rate = f'{run.passed / run.total_cases * 100:.1f}%' if run.total_cases > 0 else 'N/A'
    duration = f'{run.duration:.1f}s' if run.duration else 'N/A'
    summary_data = [
        ['Metric', 'Value'],
        ['Total Cases', str(run.total_cases)],
        ['Passed', str(run.passed)],
        ['Failed', str(run.failed)],
        ['Skipped', str(run.skipped or 0)],
        ['Error', str(run.error or 0)],
        ['Pass Rate', pass_rate],
        ['Duration', duration],
        ['Start Time', run.started_at.strftime('%Y-%m-%d %H:%M:%S') if run.started_at else 'N/A'],
        ['End Time', run.finished_at.strftime('%Y-%m-%d %H:%M:%S') if run.finished_at else 'N/A'],
    ]
    elements.append(_build_table(summary_data, col_widths=[120, 300], has_header=True))
    elements.append(Spacer(1, 6 * mm))

    # 失败用例详情
    failed_cases = [r for r in (run.results or []) if isinstance(r, dict) and r.get('status') == 'failed']
    if failed_cases:
        elements.append(Paragraph('Failed Cases', styles['Heading2']))
        failed_data = [['Name', 'Status', 'Error']]
        for case in failed_cases[:20]:  # 最多展示 20 条
            failed_data.append([
                str(case.get('name', ''))[:40],
                str(case.get('status', '')),
                str(case.get('error', ''))[:60],
            ])
        elements.append(_build_table(failed_data, col_widths=[120, 60, 240], has_header=True))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_enhanced_excel(
    test_run_id: int = None,
    project_id: int = None,
    days: int = None,
    test_type: str = None,
) -> Optional[bytes]:
    """
    生成增强版 Excel 报告（多 Sheet）

    Args:
        test_run_id: 指定单个 TestRun（与 project_id 互斥）
        project_id: 指定项目（导出项目下所有记录）
        days: 时间范围天数
        test_type: 测试类型过滤

    Returns:
        Excel 文件的 bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl 未安装，无法生成 Excel")
        return None

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')

    # 获取数据
    runs = _query_runs(test_run_id, project_id, days, test_type)

    # Sheet 1: 汇总
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _write_summary_sheet(ws_summary, runs, header_font, header_fill, header_align)

    # Sheet 2: 用例详情
    ws_cases = wb.create_sheet('Case Details')
    _write_case_details_sheet(ws_cases, runs, header_font, header_fill, header_align)

    # Sheet 3: 性能指标
    ws_perf = wb.create_sheet('Performance')
    _write_performance_sheet(ws_perf, runs, header_font, header_fill, header_align)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ── 内部工具 ──────────────────────────────────────────────────────────────────

def _query_runs(test_run_id, project_id, days, test_type) -> list:
    """查询 TestRun 列表"""
    if test_run_id:
        run = TestRun.query.get(test_run_id)
        return [run] if run else []

    query = TestRun.query
    if project_id:
        query = query.filter_by(project_id=project_id)
    if test_type:
        query = query.filter_by(test_type=test_type)
    if days:
        since = datetime.utcnow() - timedelta(days=days)
        query = query.filter(TestRun.created_at >= since)

    return query.order_by(TestRun.created_at.desc()).limit(500).all()


def _write_summary_sheet(ws, runs, header_font, header_fill, header_align):
    """写入汇总 Sheet"""
    headers = ['Run ID', 'Project', 'Test Type', 'Status', 'Total', 'Passed',
               'Failed', 'Pass Rate', 'Duration(s)', 'Start Time', 'End Time']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row, run in enumerate(runs, 2):
        ws.cell(row=row, column=1, value=run.id)
        ws.cell(row=row, column=2, value=run.project_id)
        ws.cell(row=row, column=3, value=run.test_type)
        ws.cell(row=row, column=4, value=run.status)
        ws.cell(row=row, column=5, value=run.total_cases)
        ws.cell(row=row, column=6, value=run.passed)
        ws.cell(row=row, column=7, value=run.failed)
        rate = f'{run.passed / run.total_cases * 100:.1f}%' if run.total_cases > 0 else ''
        ws.cell(row=row, column=8, value=rate)
        ws.cell(row=row, column=9, value=run.duration)
        ws.cell(row=row, column=10, value=run.started_at.strftime('%Y-%m-%d %H:%M') if run.started_at else '')
        ws.cell(row=row, column=11, value=run.finished_at.strftime('%Y-%m-%d %H:%M') if run.finished_at else '')

    for col_letter in 'ABCDEFGHIJK':
        ws.column_dimensions[col_letter].width = 14


def _write_case_details_sheet(ws, runs, header_font, header_fill, header_align):
    """写入用例详情 Sheet"""
    headers = ['Run ID', 'Case Name', 'Status', 'Duration(ms)', 'Error']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    row = 2
    for run in runs:
        for result in (run.results or []):
            if isinstance(result, dict):
                ws.cell(row=row, column=1, value=run.id)
                ws.cell(row=row, column=2, value=result.get('name', ''))
                ws.cell(row=row, column=3, value=result.get('status', ''))
                ws.cell(row=row, column=4, value=result.get('duration', ''))
                ws.cell(row=row, column=5, value=str(result.get('error', ''))[:200])
                row += 1

    for col_letter in 'ABCDE':
        ws.column_dimensions[col_letter].width = 20


def _write_performance_sheet(ws, runs, header_font, header_fill, header_align):
    """写入性能指标 Sheet"""
    perf_runs = [r for r in runs if r.test_type == 'performance']
    headers = ['Run ID', 'Status', 'Total Requests', 'Avg Response(ms)',
               'P95(ms)', 'P99(ms)', 'Error Rate', 'Throughput(rps)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row, run in enumerate(perf_runs, 2):
        ws.cell(row=row, column=1, value=run.id)
        ws.cell(row=row, column=2, value=run.status)
        # 从 results 中提取性能指标（如有）
        results = run.results or []
        if isinstance(results, list) and results:
            first = results[0] if isinstance(results[0], dict) else {}
            ws.cell(row=row, column=3, value=first.get('total_requests', ''))
            ws.cell(row=row, column=4, value=first.get('avg_response_time', ''))
            ws.cell(row=row, column=5, value=first.get('p95', ''))
            ws.cell(row=row, column=6, value=first.get('p99', ''))
            ws.cell(row=row, column=7, value=first.get('error_rate', ''))
            ws.cell(row=row, column=8, value=first.get('throughput', ''))

    for col_letter in 'ABCDEFGH':
        ws.column_dimensions[col_letter].width = 16


def _build_table(data: list, col_widths: list = None, has_header: bool = False):
    """构建 ReportLab Table"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    table = Table(data, colWidths=col_widths)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    if has_header:
        style_cmds.append(('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')))
    table.setStyle(TableStyle(style_cmds))
    return table


def _try_register_chinese_font():
    """尝试注册中文字体（可选，失败不影响）"""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        # 常见中文字体路径
        font_paths = [
            'C:/Windows/Fonts/msyh.ttc',
            'C:/Windows/Fonts/simsun.ttc',
            '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
            '/System/Library/Fonts/PingFang.ttc',
        ]
        for path in font_paths:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont('Chinese', path))
                break
    except Exception:
        pass