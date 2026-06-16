"""
导入导出服务

支持：
- Postman Collection JSON 导入
- CSV 格式批量导入用例
- Excel 格式导出测试报告

依赖：
- csv（标准库）— CSV 处理
- openpyxl — Excel 处理（可选，未安装时降级）
"""
import csv
import io
import json
import os
from datetime import datetime
from typing import Optional
from ..extensions import db
from ..models.api_test_case import ApiTestCase
from ..models.test_run import TestRun
from ..models.test_report import TestReport
from ..utils.exceptions import ValidationError, NotFoundError
from ..core.logging import get_logger

logger = get_logger(__name__)

# CSV 导入列定义
CSV_COLUMNS = ['name', 'method', 'url', 'headers', 'body', 'expected_status']


def import_from_postman_json(user_id: int, project_id: int, json_content: str,
                             collection_id: int = None) -> dict:
    """
    从 Postman Collection JSON 导入用例

    支持 Postman Collection v2.1 格式。

    Args:
        user_id: 用户 ID
        project_id: 项目 ID
        json_content: JSON 字符串
        collection_id: 目标集合 ID（可选）

    Returns:
        {total, imported, skipped, errors: [{name, reason}]}
    """
    try:
        data = json.loads(json_content)
    except json.JSONDecodeError as e:
        raise ValidationError(f"JSON 格式错误: {str(e)}")

    items = data.get('item', [])
    if not items:
        raise ValidationError("Postman Collection 中没有请求项")

    results = {'total': 0, 'imported': 0, 'skipped': 0, 'errors': []}

    def _process_items(items, parent_path=''):
        for item in items:
            # 嵌套文件夹
            if 'item' in item:
                folder_name = item.get('name', 'Unnamed')
                _process_items(item['item'], parent_path=f'{parent_path}/{folder_name}')
                continue

            results['total'] += 1
            request = item.get('request', {})
            name = item.get('name', 'Unnamed Request')

            # 解析 HTTP 方法
            method = 'GET'
            if isinstance(request, dict):
                method = request.get('method', 'GET').upper()

            # 解析 URL
            url = ''
            if isinstance(request, dict):
                url_obj = request.get('url', {})
                if isinstance(url_obj, str):
                    url = url_obj
                elif isinstance(url_obj, dict):
                    raw = url_obj.get('raw', '')
                    url = raw

            if not url:
                results['skipped'] += 1
                results['errors'].append({'name': name, 'reason': 'URL 为空'})
                continue

            # 解析 Headers
            headers = {}
            if isinstance(request, dict):
                for h in request.get('header', []):
                    if isinstance(h, dict) and not h.get('disabled'):
                        headers[h.get('key', '')] = h.get('value', '')

            # 解析 Body
            body = None
            if isinstance(request, dict):
                body_obj = request.get('body', {})
                if isinstance(body_obj, dict):
                    mode = body_obj.get('mode', '')
                    if mode == 'raw':
                        raw = body_obj.get('raw', '')
                        try:
                            body = json.loads(raw)
                        except (json.JSONDecodeError, TypeError):
                            body = raw
                    elif mode == 'urlencoded':
                        body = {p.get('key', ''): p.get('value', '') for p in body_obj.get('urlencoded', []) if isinstance(p, dict)}
                    elif mode == 'formdata':
                        body = {p.get('key', ''): p.get('value', '') for p in body_obj.get('formdata', []) if isinstance(p, dict)}

            # 检查是否重复（同名同 URL）
            existing = ApiTestCase.query.filter_by(
                user_id=user_id, project_id=project_id, name=name, url=url,
            ).first()
            if existing:
                results['skipped'] += 1
                results['errors'].append({'name': name, 'reason': '已存在同名同 URL 用例'})
                continue

            # 创建用例
            case = ApiTestCase(
                user_id=user_id,
                project_id=project_id,
                collection_id=collection_id,
                name=name,
                description=f'从 Postman 导入 ({parent_path})' if parent_path else '从 Postman 导入',
                method=method,
                url=url,
                headers=headers,
                body=body,
                body_type='json',
            )
            db.session.add(case)
            results['imported'] += 1

    _process_items(items)

    db.session.commit()
    logger.info("Postman 导入完成", user_id=user_id, project_id=project_id,
                total=results['total'], imported=results['imported'])
    return results


def import_from_csv(user_id: int, project_id: int, csv_content: str,
                    collection_id: int = None) -> dict:
    """
    从 CSV 格式批量导入用例

    CSV 列：name, method, url, headers, body, expected_status

    Args:
        user_id: 用户 ID
        project_id: 项目 ID
        csv_content: CSV 文本内容
        collection_id: 目标集合 ID（可选）

    Returns:
        {total, imported, skipped, errors: [{row, reason}]}
    """
    reader = csv.DictReader(io.StringIO(csv_content))

    # 校验列头
    if reader.fieldnames is None:
        raise ValidationError("CSV 文件为空")
    required_cols = {'name', 'method', 'url'}
    missing = required_cols - set(reader.fieldnames)
    if missing:
        raise ValidationError(f"CSV 缺少必需列: {', '.join(missing)}")

    results = {'total': 0, 'imported': 0, 'skipped': 0, 'errors': []}

    for row_num, row in enumerate(reader, start=2):
        results['total'] += 1

        name = row.get('name', '').strip()
        method = row.get('method', 'GET').strip().upper()
        url = row.get('url', '').strip()

        if not name or not url:
            results['skipped'] += 1
            results['errors'].append({'row': row_num, 'reason': 'name 或 url 为空'})
            continue

        # 解析 headers（JSON 字符串）
        headers = {}
        headers_str = row.get('headers', '').strip()
        if headers_str:
            try:
                headers = json.loads(headers_str)
            except json.JSONDecodeError:
                results['skipped'] += 1
                results['errors'].append({'row': row_num, 'reason': 'headers JSON 格式错误'})
                continue

        # 解析 body
        body = None
        body_str = row.get('body', '').strip()
        if body_str:
            try:
                body = json.loads(body_str)
            except (json.JSONDecodeError, TypeError):
                body = body_str

        # 去重检查
        existing = ApiTestCase.query.filter_by(
            user_id=user_id, project_id=project_id, name=name, url=url,
        ).first()
        if existing:
            results['skipped'] += 1
            results['errors'].append({'row': row_num, 'reason': '已存在同名同 URL 用例'})
            continue

        case = ApiTestCase(
            user_id=user_id,
            project_id=project_id,
            collection_id=collection_id,
            name=name,
            description='从 CSV 导入',
            method=method,
            url=url,
            headers=headers,
            body=body,
            body_type='json',
        )
        db.session.add(case)
        results['imported'] += 1

    db.session.commit()
    logger.info("CSV 导入完成", user_id=user_id, project_id=project_id,
                total=results['total'], imported=results['imported'])
    return results


def export_test_report_excel(test_run_id: int) -> Optional[bytes]:
    """
    导出测试报告为 Excel 格式

    Args:
        test_run_id: TestRun ID

    Returns:
        Excel 文件的 bytes，None 表示 openpyxl 未安装
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl 未安装，无法导出 Excel")
        return None

    run = TestRun.query.get(test_run_id)
    if not run:
        raise NotFoundError("测试执行记录", test_run_id)

    wb = Workbook()

    # Sheet 1: 汇总
    ws_summary = wb.active
    ws_summary.title = '执行摘要'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    summary_headers = ['字段', '值']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    summary_data = [
        ('执行 ID', run.id),
        ('项目 ID', run.project_id),
        ('测试类型', run.test_type),
        ('测试对象', run.test_object_name or ''),
        ('状态', run.status),
        ('总用例数', run.total_cases),
        ('通过', run.passed),
        ('失败', run.failed),
        ('跳过', run.skipped),
        ('错误', run.error),
        ('通过率', f"{run.passed / run.total_cases * 100:.1f}%" if run.total_cases > 0 else 'N/A'),
        ('耗时(秒)', run.duration or ''),
        ('开始时间', run.started_at.isoformat() if run.started_at else ''),
        ('结束时间', run.finished_at.isoformat() if run.finished_at else ''),
        ('触发方式', run.triggered_by or ''),
    ]
    for row, (field, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=field)
        ws_summary.cell(row=row, column=2, value=str(value))

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 30

    # Sheet 2: 用例详情
    ws_cases = wb.create_sheet('用例详情')
    case_headers = ['用例名称', '状态', '耗时(秒)', '错误信息']
    for col, header in enumerate(case_headers, 1):
        cell = ws_cases.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    results = run.results or []
    for row, result in enumerate(results, 2):
        if isinstance(result, dict):
            ws_cases.cell(row=row, column=1, value=result.get('name', ''))
            ws_cases.cell(row=row, column=2, value=result.get('status', ''))
            ws_cases.cell(row=row, column=3, value=result.get('duration', ''))
            ws_cases.cell(row=row, column=4, value=result.get('error', ''))

    for col_letter in ['A', 'B', 'C', 'D']:
        ws_cases.column_dimensions[col_letter].width = 25

    # 保存到 bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_test_report_csv(test_run_id: int) -> str:
    """
    导出测试报告为 CSV 格式

    Args:
        test_run_id: TestRun ID

    Returns:
        CSV 文本内容
    """
    run = TestRun.query.get(test_run_id)
    if not run:
        raise NotFoundError("测试执行记录", test_run_id)

    output = io.StringIO()
    writer = csv.writer(output)

    # 写入汇总信息
    writer.writerow(['字段', '值'])
    writer.writerow(['执行 ID', run.id])
    writer.writerow(['测试类型', run.test_type])
    writer.writerow(['状态', run.status])
    writer.writerow(['总用例数', run.total_cases])
    writer.writerow(['通过', run.passed])
    writer.writerow(['失败', run.failed])
    writer.writerow(['通过率', f"{run.passed / run.total_cases * 100:.1f}%" if run.total_cases > 0 else 'N/A'])
    writer.writerow([])

    # 写入用例详情
    writer.writerow(['用例名称', '状态', '耗时(秒)', '错误信息'])
    results = run.results or []
    for result in results:
        if isinstance(result, dict):
            writer.writerow([
                result.get('name', ''),
                result.get('status', ''),
                result.get('duration', ''),
                result.get('error', ''),
            ])

    return output.getvalue()


def generate_csv_template() -> str:
    """生成 CSV 导入模板"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_COLUMNS)
    writer.writerow(['用户登录接口', 'POST', 'https://api.example.com/login', '{"Content-Type": "application/json"}', '{"username": "test", "password": "123456"}', '200'])
    writer.writerow(['获取用户信息', 'GET', 'https://api.example.com/users/1', '{}', '', '200'])
    return output.getvalue()